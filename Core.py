#author: sam gates

#this is a program built to grade excel papers
#it is fully featured with its own grammar to easily
#write new assignments and have them graded in a jiffy!

from openpyxl import Workbook
from openpyxl import cell
from openpyxl import load_workbook
import re
import os
import sys
import math

class AutoGrader:

    def main(self):

        #ask for assignment folder containing the things to grade
        self.Assignment = input('Enter Assignment Folder: ')

        #ask for the .sag file that we will parse to get our answer key
        self.keyFile = input('Enter key file name: ') + '.sag'

        self.sheetToGradeNum = 0
        self.sheetToGradeNumbers=[]
        self.aSyntax=[]
        self.readAssignmentKey(self.keyFile)
        self.pointLoss = 0
        for x in range(0,self.aSyntax.__len__()):
            for i in range(0, self.aSyntax[x].__len__()):
                print(self.aSyntax[x][i].__str__() + "\n")

        #gF = Graded File, the file we write the output to
        self.gF = open(self.Assignment+'Graded.txt','w')


        #load in each paper and grade them!
        for file in os.listdir(self.Assignment):
            if file.endswith(".xlsx"):
                print(self.Assignment+'/'+file.__str__())
                try:
                    self.gradePaper(self.Assignment+'/'+file.__str__())
                except:
                    print(file.__str__()+' had an error, OOPS:',sys.exc_info()[0])
                    self.gF.write('error\n\n')

        self.gF.close()

    def gradePaper(self,path):

        #loading in the excel sheet
        fileName = path.replace(self.Assignment+'/','')
        fileName = fileName.replace('.xlsx','')
        self.gF.write(fileName+'\n\n')
 #       print("load wb")
        awb = load_workbook(path)
  #      print("load wb2")
        awb2=load_workbook(path, data_only=True)
   #     print("done loading wb and wb2")
        asheets = []
        asheetsNotFormulas=[]
        for s in awb._sheets:
            asheets.append(s)
        for s in awb2._sheets:
            asheetsNotFormulas.append(s)
        print('Start')

        #If we have at least one sheet, we drop info the grading loop
        if asheets.__len__()>0:
            #self.gF.write('Number of Sheets: ' + asheets.__len__().__str__()+'\n')
            score = 60
 #           print("score:" + score.__str__())
            try:
                #for each sheet
                for sheetNum in range(0,self.sheetToGradeNumbers.__len__()):
                    self.sheetToGradeNum=self.sheetToGradeNumbers[sheetNum]
                    ws=asheets[self.sheetToGradeNum]

                    ws2=asheetsNotFormulas[self.sheetToGradeNum]
                    #Go through each question
                    for qNum in range(0,self.aSyntax[sheetNum].__len__()):
                        #print("Question Number: "+ qNum.__str__())
                        curQ = self.aSyntax[sheetNum][qNum]
#                        print("question: ",curQ)
                        #Check to see if it is a single-condition question, check count if so
                        if curQ.__len__()==1:
                            #print("ayy")
                            if not self.checkStatement(curQ[0],ws,ws2,score,True):
                                score-=self.pointLoss
                                print("Losing " + self.pointLoss.__str__()+" points, now at "+score.__str__() + " on question number "+qNum.__str__())
                                #print("false")

                            #print('done')

                        elif curQ.__len__()>1:
                            #check each condition, if any fail then we break out of loop and subtract points
                            for i in range(0,curQ.__len__()):
            #                    print("Current Q: " + curQ[i].__str__())
                                finalCondition = (i==curQ.__len__()-1)
                                if not self.checkStatement(curQ[i],ws,ws2,score,finalCondition):
                                    #print('False')
                                    score-=self.pointLoss
                                    print("Losing " + self.pointLoss.__str__()+" points, now at "+score.__str__() + " on question number "+qNum.__str__())
                                    break
                                else:
                                    continue
#                                    print('True')

            except AttributeError as e:
                self.gF.write('Student left cells blank\n\n\n')
                print('Student left cells blank: '+str(e))

            print('~~~~~~~~~~~~~~')
            print('Score: ' + score.__str__())
            print('~~~~~~~~~~~~~~')
            self.gF.write('Score: '+score.__str__()+'\n\n\n')

    #Check for Statement stmt in worksheet ws
    def checkStatement(self, stmt, ws, ws2, score,finalCondition):
        #print("Checking Statement")
        #parse stmt into useful information
        cellToCheck=stmt[0]
        n=stmt[1]
        valToCheck=stmt[2]
        comment=stmt[3]
        pointVal=stmt[4]
        correctAnswer = stmt[5]

        self.pointLoss=pointVal
        #print("Comment:" +comment)

        #check the workbook for the desired statement count, fail and subtract score if is less
        if(n > 0):
            if valToCheck.upper()=="ANS":
                self.gF.write(comment+'\n')
                print(comment)
                #print("Statement less than desired")
                return False
            elif valToCheck.upper()!="XXX" and (str(ws[cellToCheck].value)).upper().count(valToCheck.upper())<n:
                #print(ws[cellToCheck].value)
                #print(ws.title)
                self.gF.write(comment+'\n')
                print(comment)
                #print("Statement less than desired")
                return False
            elif finalCondition and correctAnswer!='XX' and not self.isFloat(ws2[cellToCheck].value) and str(ws2[cellToCheck].value).upper()!=correctAnswer.upper():
                self.gF.write('Answer did not match correct value in cell '+cellToCheck+' on sheet ' + ws.title.__str__()+' but used the correct formulas (-1pt)\n')
                print('Answer Wrong Not Decimal')
                #stmt[4]=1

                self.pointLoss = 1
                #offsetting the -5 from getting this wrong so its only -1... shuddup
                return False
            elif finalCondition and correctAnswer!='XX' and self.isFloat(ws2[cellToCheck].value):
                #if its a decimal we want to shave it off and check the first 6 decimal places
                ws2float = round(float(ws2[cellToCheck].value),2)
                correctFloat = round(float(correctAnswer),2)
    #            print(correctFloat.__str__()+" and ws2 is "+ws2float.__str__())
                if ws2float!=correctFloat:
                    self.gF.write('Answer did not match correct value in cell '+cellToCheck+' on sheet ' + ws.title.__str__()+' but used the correct formulas (-1pt)\n')
                    #stmt[4]=1
                    #offsetting the -5 from getting this wrong so its only -1... shuddup
                    self.pointLoss = 1

                    print(correctFloat.__str__()+" and ws2 is "+ws2float.__str__() )
                    #print('Answer Wrong')
                    return False
                else:
                    #print("Answer Correct")
                    return True
            else:
                #print("Answer Correct")
                return True
        else:
            if str(ws[cellToCheck].value).upper().count(valToCheck)>=-n:

                self.gF.write(comment+'\n')
                print(comment)
                #print("Statement less than desired")
                return False
            else:
                return True

    #Read in an assignment key to the self.aSyntax variable
    def readAssignmentKey(self,keyPath):
        aKey=open(keyPath,'r')

        #Break down lines in the key we just opened
        lines = [line.rstrip('\n') for line in aKey]

        curSheet = 0
        state = 0

        newQuestion = []

        multConditions = False

        for i in range(0,lines.__len__()):

            # Get the current line we are parsing
            l = lines[i]

            #State 0 = Start new Sheet
            #############################################################
            if state == 0:
                if l[0] == '#':
                    # Skip Line, continue because this is a comment
                    continue
                elif l[0] == '=':
                    self.sheetToGradeNum=curSheet
                    self.sheetToGradeNumbers.append(int(l[1]))
                    #curSheet =0
                    self.aSyntax.append([])
                    state = 1

            #State 1 = Check for New Question
            #############################################################
            elif state == 1:
                if l[0] == '*':
                    #print('Start Multiple Condition Statement')
                    newQuestion = []
                    state = 3
                elif l[0] == '[':
                    newQuestion = []
                    state = 2
                elif l[0] == '#':
                    continue
                elif l[0]=="=":
                    curSheet=curSheet+1
                    self.sheetToGradeNum=int(l[1])
                    self.sheetToGradeNumbers.append(int(l[1]))
                    #curSheet =0
                    self.aSyntax.append([])
                    state = 1
                else:
                    print('Error on line '+i.__str__()+'), stuck in state 1')

            #State 2 = Read single statement in, fall back to S1
            #############################################################
            if state == 2:

                #parse into tokens separated by spaces
                parsedLine = re.split('_',l)

                newStatement = self.readStatement(parsedLine)
                newQuestion=[newStatement]
                self.aSyntax[curSheet].append(newQuestion)
                state = 1

            #State 3 = Read multiple statements in as a multi-condition, fall back to S1 when Question ends
            #############################################################
            elif state == 3:
                parsedLine = re.split('_',l)

                #Start of multi-condition
                if parsedLine[0].startswith('*['):
                    if not multConditions:
                        multConditions = True
                        parsedLine[0]=parsedLine[0].replace('*','')
                        newQuestion.append(self.readStatement(parsedLine))

                    #if multConditions, we need to finish the last question and then start a new one
                    #as this is the start of a new multi-condition question
                    elif multConditions:
                        self.aSyntax[curSheet].append(newQuestion)
                        newQuestion = []
                        parsedLine[0]=parsedLine[0].replace('*','')
                        newQuestion.append(self.readStatement(parsedLine))

                #continuation of multi-condition
                elif parsedLine[0].startswith('**[') and multConditions:
                    parsedLine[0]=parsedLine[0].replace('**','')
                    newQuestion.append(self.readStatement(parsedLine))

                #fall out of multi-condition when the line starts with [
                elif parsedLine[0].startswith('[') and multConditions:
                    #print('End Multiple Condition Statement')
                    multConditions = False
                    self.aSyntax[curSheet].append(newQuestion)

                    #ensure that the line we are currently on gets read
                    newQuestion = [self.readStatement(parsedLine)]
                    self.aSyntax[curSheet].append(newQuestion)
                    state = 1

            #if a multi-condition is the last line of the assignment file
            if i == lines.__len__()-1 and multConditions:
                self.aSyntax[curSheet].append(newQuestion)


    #Read a statement from the parsed line in readAssignmentKey
    def readStatement(self,parsedLine):

        #This is the structure of our statements
        #[CellToCheck, operator, ValueToCheckFor, Grader Comment, PointValue, Correct Answer]

        newStatement = ['', #Cell To Check
                        0,  #Operator
                        '', #Value we want to check is in Cell To Check
                        '', #Grader's comment about getting this question condition wrong
                        0,  #The point value that this question condition is worth
                        ''] #The correct answer to compare against
        newStatement[0] = parsedLine[0][1:-1]
        #print(newStatement[0])

        ####Check Operator here####
        #if Check keyword is used, we search for soemthing that will result in a correct answer
        if 'Check' in parsedLine[1]:
            newStatement[1] = int(parsedLine[1][-1])
        #if Discard keyword is used, we search for something that will result in a wrong answer
        elif 'Discard' in parsedLine[1]:
            newStatement[1]=-1*int(parsedLine[1][-1])
        else:
            print(parsedLine[1])

        #Value To Check For
        newStatement[2] = parsedLine[2][1:-1]

        #Point Value
        newStatement [4] = int(parsedLine[4][1:])

        #the grader's comment
        commentString = parsedLine[3]+' (-'+newStatement[4].__str__()+'pts)'
        newStatement[3] = commentString

        #correct answer
        newStatement[5] = parsedLine[5]

        return newStatement

    def isFloat(self,value):
        if str(value).upper()=="TRUE" or str(value).upper()=="FALSE":
            return False
        try:
            float(value)
            return True
        except ValueError:
            return False


AG  = AutoGrader()

AG.main()